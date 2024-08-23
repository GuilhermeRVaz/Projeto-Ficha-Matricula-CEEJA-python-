import os
import io
import tkinter as tk
from tkinter import messagebox
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import warnings
import time
import subprocess
import json
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Configurações
SCOPES = ['https://www.googleapis.com/auth/drive.readonly']
SPREADSHEET_IDS = {
    'F': '10_beB24nkw37BsWQaCrS2lrFspcAarA1',
    'M': '1hGsycXU1JM-saNrluhbHumtsxj6vMRFm'
}
RANGES = {
    'F': [
        '2024_Matriculas',
        'EF_Geral_Alfabetica_31122023'
    ],
    'M': [
        '2024_MATRICULAS',
        'EM_Geral_Alfabetica_31122023'
    ]
}

def authenticate():
    creds = None
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        with open('token.json', 'w') as token:
            token.write(creds.to_json())
    return creds

def download_file(service, file_id):
    request = service.files().get_media(fileId=file_id)
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while not done:
        status, done = downloader.next_chunk()
        print(f"Download {int(status.progress() * 100)}%.")
    fh.seek(0)
    return fh

def search_passport(passport, service):
    prefix = passport[0]  # 'F' ou 'M'
    file_id = SPREADSHEET_IDS.get(prefix)
    sheet_names = RANGES.get(prefix)

    if not file_id or not sheet_names:
        return "Planilha não encontrada para o formato do passaporte."

    file_handle = download_file(service, file_id)
    
    for sheet_name in sheet_names:
        print(f"Procurando na aba {sheet_name}...")
        # Ignore warnings about date parsing issues
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            df = pd.read_excel(file_handle, sheet_name=sheet_name, engine='openpyxl')

        if 'Unnamed: 1' in df.columns:
            passaporte_col = df.columns.get_loc('Unnamed: 1')
            nome_col = df.columns.get_loc('Unnamed: 2')
            ra_col = df.columns.get_loc('Unnamed: 3')
            rg_col = df.columns.get_loc('Unnamed: 4')
            telefone_col = df.columns.get_loc('Unnamed: 5')
            cidade_col = df.columns.get_loc('Unnamed: 6')
            serie_concluida_col = df.columns.get_loc('Unnamed: 9')
            serie_matricular_col = df.columns.get_loc('Unnamed: 10')

            if prefix == 'M':
                prof_disciplina_col = df.columns.get_loc('Unnamed: 23')
                observacao_col = df.columns.get_loc('Unnamed: 34')
            else:  # prefix == 'F'
                prof_disciplina_col = df.columns.get_loc('Unnamed: 18')
                observacao_col = df.columns.get_loc('Unnamed: 25')

            # Limpeza e comparação dos dados
            df.iloc[:, passaporte_col] = df.iloc[:, passaporte_col].astype(str).str.strip()  # Remove espaços extras
            result_row = df[df.iloc[:, passaporte_col] == passport]
            if not result_row.empty:
                # Formata os resultados para exibição
                result = {
                    'Passaporte': result_row.iloc[0, passaporte_col],
                    'Nome': result_row.iloc[0, nome_col],
                    'RA': result_row.iloc[0, ra_col],
                    'RG': result_row.iloc[0, rg_col],
                    'Telefone': result_row.iloc[0, telefone_col],
                    'Cidade': result_row.iloc[0, cidade_col],
                    'Série Concluída': result_row.iloc[0, serie_concluida_col],
                    'Série a ser Matriculado': result_row.iloc[0, serie_matricular_col],
                    'Prof. 1ª Disciplina': result_row.iloc[0, prof_disciplina_col],
                    'Observação': result_row.iloc[0, observacao_col]
                }
                return result

    return "Passaporte não encontrado."

def extract_site_data(ra_number):
    driver_path = 'C:/Users/e497976a/Documents/aut_passap_plani/chromedriver.exe'
    service = Service(executable_path=driver_path)
    driver = webdriver.Chrome(service=service)

    # Carregar dados de login do arquivo JSON
    with open('login_credentials.json') as f:
        credentials = json.load(f)
    username = credentials['username']
    password = credentials['password']
    
    driver.get('https://sed.educacao.sp.gov.br/NCA/FichaAluno/Index')
    time.sleep(3)

    login_field = driver.find_element(By.XPATH, '//*[@id="name"]')
    login_field.send_keys(username)
    password_field = driver.find_element(By.XPATH, '//*[@id="senha"]')
    password_field.send_keys(password)
    time.sleep(3)
    login_button = driver.find_element(By.XPATH, '//*[@id="botaoEntrar"]')
    login_button.click()
    time.sleep(3)
    search_box = driver.find_element(By.XPATH, '//*[@id="decorMenuFilterTxt"]')
    search_box.send_keys('ficha do aluno')
    search_box.send_keys(Keys.ENTER)
    time.sleep(3)
    
    wait = WebDriverWait(driver, 10)
    dropdown_menu = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="Pesquisa"]/div[1]/div[2]/div/div/button/div/div/div')))
    dropdown_menu.click()
    ra_option = wait.until(EC.visibility_of_element_located((By.XPATH, '//*[@id="bs-select-1-1"]/span')))
    ra_option.click()
    search_box_RAaluno = driver.find_element(By.XPATH, '//*[@id="txtRa"]')

    # Converte ra_number para string e remove '-' e o último dígito
    clean_ra_number = str(ra_number).replace('-', '')[:-1] 
    search_box_RAaluno.send_keys(clean_ra_number)
    search_box_Pesquisar = driver.find_element(By.XPATH, '//*[@id="btnPesquisar"]')
    search_box_Pesquisar.click()
    time.sleep(3)
    
    try:
        search_botao_editar = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'i.icone-tabela-editar[title="Editar"]')))
        search_botao_editar.click()
    except Exception as e:
        print(f"Erro ao encontrar o botão 'Editar': {e}")

    time.sleep(3)
    
    site_data = {}
    try:
        endereco_element = driver.find_element(By.XPATH, '//*[@id="Endereco"]')
        site_data['Endereco'] = endereco_element.get_attribute('value')
    except Exception as e:
        print(f'Erro ao encontrar o campo de endereço: {e}')
    try:
        endereco_nr_element = driver.find_element(By.XPATH, '//*[@id="EnderecoNR"]')
        site_data['EnderecoNR'] = endereco_nr_element.get_attribute('value')
    except Exception as e:
        print(f'Erro ao encontrar o número da casa: {e}')
    try:
        cep_element = driver.find_element(By.CSS_SELECTOR, 'input.form-control.cep')
        site_data['CEP'] = cep_element.get_attribute('value')
    except Exception as e:
        print(f'Erro ao encontrar o campo de CEP: {e}')
    try:
        dt_nascimento_element = driver.find_element(By.XPATH, '//*[@id="DtNascimento"]')
        site_data['Data de Nascimento'] = dt_nascimento_element.get_attribute('value')
    except Exception as e:
        print(f'Erro ao encontrar o campo de Data de Nascimento: {e}')
    
    driver.quit()
    return site_data

def update_excel(result, site_data):
    # Abre o arquivo Excel
    wb = load_workbook('Passaporte_Geral_2024.xlsx')
    ws = wb.active

    # Atualizar com dados da pesquisa
    ws['K3'] = result.get('Passaporte', '')
    ws['C3'] = result.get('Nome', '')
    ws['J4'] = result.get('RA', '')
    ws['G4'] = result.get('RG', '')
    ws['C6'] = result.get('Telefone', '')
    ws['I5'] = result.get('Cidade', '')
    ws['D25'] = result.get('Série Concluída', '')
    ws['G25'] = result.get('Série a ser Matriculado', '')
    ws['D26'] = result.get('Observação', '')

    # Atualizar com dados extraídos do site
    if site_data:
        endereco_completo = f"{site_data.get('Endereco', '')}, {site_data.get('EnderecoNR', '')}"
        ws['C5'] = endereco_completo
        ws['L5'] = site_data.get('CEP', '')
        ws['D4'] = site_data.get('Data de Nascimento', '')

    wb.save('Passaporte_Geral_2024.xlsx')

    # Abrir o arquivo em tela
    try:
        if os.name == 'nt':  # Windows
            os.startfile('Passaporte_Geral_2024.xlsx')
        elif os.name == 'posix':  # macOS ou Linux
            subprocess.call(('open', 'Passaporte_Geral_2024.xlsx'))
    except Exception as e:
        print(f"Erro ao abrir o arquivo: {e}")

import tkinter as tk
from tkinter import messagebox, PhotoImage
from openpyxl import load_workbook
import os


def clear_form(ws):
    # Limpar campos específicos
    ws['C6'] = ''
    ws['B7'] = ''
    ws['F7'] = ''
    ws['M7'] = ''
    ws['C8'] = ''

    # Gêmeos
    ws['O9'] = 'Sim( )'
    ws['P9'] = 'Não( )'

    # Limpar cor/raça
    ws['G8'] = '( )'
    ws['I8'] = '( )'
    ws['K8'] = '( )'
    ws['M8'] = '( )'
    ws['O8'] = '( )'
    ws['Q8'] = '( )'

    ws['G9'] = ''
    ws['C9'] = ''
    ws['C10'] = ''
    ws['G10'] = ''
    ws['O10'] = ''
    ws['Q10'] = ''

    ws['J14'] = ''
    ws['K13'] = '( )'
    ws['Q13'] = '( )'

    ws['B19'] = ''
    ws['O19'] = ''
    ws['C20'] = ''
    ws['M20'] = '( )'
    ws['O20'] = '( )'

    ws['B21'] = ''
    ws['F21'] = ''
    ws['M21'] = ''
    ws['F22'] = ''
    ws['M22'] = ''

    ws['I25'] = '( )'
    ws['K25'] = '( )'
    ws['M25'] = '( )'
    ws['O25'] = '( )'
    ws['Q25'] = '( )'

    ws['I26'] = '( )'
    ws['K26'] = '( )'
    ws['M26'] = '( )'
    ws['O26'] = '( )'

    ws['J30'] = '( )'
    ws['L30'] = '( )'

    ws['K15'] = '( )'
    ws['M15'] = '( )'
    ws['M16'] = '( )'
    ws['K17'] = '( )'
    ws['M17'] = '( )'

    ws['I31'] = 'Não( )'
    ws['I32'] = 'Sim( )'
    ws['I33'] = '( )'
    ws['I34'] = '( )'
    ws['I35'] = '( )'
    ws['I37'] = '( )'
    ws['I38'] = '( )'

    # Limpar campos de documentos entregues
    ws['A40'] = '( )'
    ws['A41'] = '( )'
    ws['A42'] = '( )'
    ws['A43'] = '( )'
    ws['D40'] = '( )'
    ws['D41'] = '( )'
    ws['D42'] = '( )'
    ws['D43'] = '( )'
    ws['I40'] = '( )'
    ws['I41'] = '( )'
    ws['I42'] = '( )'
    ws['M40'] = '( )'
    ws['M41'] = '( )'
    ws['M42'] = '( )'
    ws['A45'] = '( )'

def populate_form(form_data):
    # Carregar a planilha Excel
    wb = load_workbook('FICHA_DE_MATRÍCULA_2024.xlsx')
    ws = wb.active

    clear_form(ws)

    # Preencher as células com base nos dados do formulário
    ws['C6'] = form_data.get('Nome', '')
    ws['B7'] = form_data.get('RG', '')
    ws['F7'] = form_data.get('CPF', '')
    ws['M7'] = form_data.get('RA', '')
    ws['C8'] = form_data.get('Estado Civil', '')

    # Cor/raça
    cor_raca = form_data.get('Cor/raça', '')
    ws['G8'] = '(X)' if cor_raca == 'Branco' else '( )'
    ws['I8'] = '(X)' if cor_raca == 'Preto' else '( )'
    ws['K8'] = '(X)' if cor_raca == 'Pardo' else '( )'
    ws['M8'] = '(X)' if cor_raca == 'Amarelo' else '( )'
    ws['O8'] = '(X)' if cor_raca == 'Indígena' else '( )'
    ws['Q8'] = '(X)' if cor_raca == 'Outra' else '( )'

    # Outras informações do formulário
    ws['G9'] = form_data.get('Nome da Mãe', '')

    # Gêmeo
    gemeo_sim = form_data.get('Gêmeo', '( )')
    ws['O9'] = '(X)Sim' if gemeo_sim == "Sim" else '( )'
    ws['P9'] = '(X)Não' if gemeo_sim == "Não" else '( )'

    ws['C10'] = form_data.get('Nascimento', '')
    ws['G10'] = form_data.get('Município', '')
    ws['O10'] = form_data.get('UF', '')
    ws['Q10'] = form_data.get('País', '')

    # Opção de Itinerário
    if form_data.get('Opção de Itinerário') == 'Ciências Naturais/Matemática':
        ws['K13'] = '(X)'
        ws['Q13'] = '( )'
    else:
        ws['Q13'] = '(X)'
        ws['K13'] = '( )'

    # Outros campos preenchidos
    ws['B19'] = form_data.get('Endereço', '')
    ws['O19'] = form_data.get('Número', '')
    ws['C20'] = form_data.get('Bairro', '')

    # Zona Urbana ou Rural
    if form_data.get('Urbana/Rural') == 'Urbana':
        ws['M20'] = '( X )'
        ws['O20'] = '(  )'
    elif form_data.get('Urbana/Rural') == 'Rural':
        ws['M20'] = '(  )'
        ws['O20'] = '( X )'
    else:
        ws['M20'] = '(  )'
        ws['O20'] = '(  )'

    ws['B21'] = form_data.get('CEP', '')
    ws['F21'] = form_data.get('Cidade', '')
    ws['M21'] = form_data.get('UF_Cidade', '')
    ws['F22'] = form_data.get('Telefone Celular', '')
    ws['M22'] = form_data.get('Telefone Recado', '')

    # Preenchimento adicional baseado no nível de ensino e série/termo
    nivel_ensino = form_data.get('Requer Matrícula no', '')

    if nivel_ensino == 'Ensino Fundamental':
        ws['I25'] = '(X)'  # Marca "Ensino Fundamental"

        termo = form_data.get('Termo/Série', '')
        if termo == '1º Termo':
            ws['K25'] = '(X)'
        elif termo == '2º Termo':
            ws['M25'] = '(X)'
        elif termo == '3º Termo':
            ws['O25'] = '(X)'
        elif termo == '4º Termo':
            ws['Q25'] = '(X)'

    elif nivel_ensino == 'Ensino Médio':
        ws['I26'] = '(X)'  # Marca "Ensino Médio"

        serie = form_data.get('Termo/Série', '')
        if serie == '1ª Série':
            ws['K26'] = '(X)'
        elif serie == '2ª Série':
            ws['M26'] = '(X)'
        elif serie == '3ª Série':
            ws['O26'] = '(X)'

    if form_data.get('Ensino Religioso') == 'Sim':
        ws['J30'] = '(X)'
    else:
        ws['L30'] = '(X)'

    # Já estudou nesta Unidade?
    if form_data.get('Estudou nesta U.E.') == 'Sim':
        ws['K15'] = '(X)'
    else:
        ws['M15'] = '(X)'

    # Aproveitamento de Estudos
    if form_data.get('Aproveitamento de Estudos') == 'Sim':
        ws['K16'] = '(X)'
    else:
        ws['M16'] = '(X)'

    # Portador de necessidades ou PCD
    if form_data.get('Portador de necessidades ou PCD') == 'Sim':
        ws['K17'] = '(X)'
    else:
        ws['M17'] = '(X)'

    # Necessidade Especial
    ws['J14'] = form_data.get('Se sim, qual', '')

    # Preencher os campos de documentos entregues
    if form_data['Doc_RG']:
        ws['A40'] = '(X)'
    if form_data['Doc_CPF']:
        ws['A41'] = '(X)'
    if form_data['Foto']:
        ws['A42'] = '(X)'
    if form_data['Requerimento de Matrícula']:
        ws['A43'] = '(X)'
    if form_data['Histórico Escolar']:
        ws['D40'] = '(X)'
    if form_data['Comprovante de Endereço']:
        ws['D41'] = '(X)'
    if form_data['Carteira de Vacinação']:
        ws['D42'] = '(X)'
    if form_data['Certidão de Nascimento']:
        ws['D43'] = '(X)'
    if form_data['Ficha de Aproveitamento de Estudos']:
        ws['I40'] = '(X)'
    if form_data['Relatório Médico']:
        ws['I41'] = '(X)'
    if form_data['Declaração de Transferência']:
        ws['I42'] = '(X)'
    if form_data['Requerimento de Transferência']:
        ws['M40'] = '(X)'
    if form_data['Declaração de Matrícula']:
        ws['M41'] = '(X)'
    if form_data['Boletim Escolar']:
        ws['M42'] = '(X)'
    if form_data['Certificado de Conclusão']:
        ws['A45'] = '(X)'

    # Salvar a planilha Excel
    wb.save('FICHA_DE_MATRÍCULA_2024.xlsx')
    messagebox.showinfo("Sucesso", "Ficha de matrícula preenchida com sucesso!")

# Função para lidar com o clique do botão "Enviar"
    
def enviar():
    # Obter os dados do formulário
    form_data = {
        'Nome': nome_entry.get(),
        'RG': rg_entry.get(),
        'CPF': cpf_entry.get(),
        'RA': ra_entry.get(),
        'Estado Civil': estado_civil_entry.get(),
        'Cor/raça': cor_var.get(),
        'Nome da Mãe': nome_mae_entry.get(),
        'Gêmeo': gemeo_var.get(),
        'Nascimento': nascimento_entry.get(),
        'Município': municipio_entry.get(),
        'UF': uf_entry.get(),
        'País': pais_entry.get(),
        'Opção de Itinerário': itinerario_var.get(),
        'Endereço': endereco_entry.get(),
        'Número': numero_entry.get(),
        'Bairro': bairro_entry.get(),
        'Urbana/Rural': zona_var.get(),
        'CEP': cep_entry.get(),
        'Cidade': cidade_entry.get(),
        'UF_Cidade': uf_cidade_entry.get(),
        'Telefone Celular': telefone_celular_entry.get(),
        'Telefone Recado': telefone_recado_entry.get(),
        'Requer Matrícula no': nivel_var.get(),
        'Termo/Série': serie_var.get(),
        'Ensino Religioso': ensino_religioso_var.get(),
        'Estudou nesta U.E.': estudou_ue_var.get(),
        'Aproveitamento de Estudos': aproveitamento_var.get(),
        'Portador de necessidades ou PCD': necessidades_var.get(),
        'Se sim, qual': necessidade_qual_entry.get(),
        'Doc_RG': doc_rg_var.get(),
        'Doc_CPF': doc_cpf_var.get(),
        'Foto': doc_foto_var.get(),
        'Requerimento de Matrícula': doc_requerimento_var.get(),
        'Histórico Escolar': doc_historico_var.get(),
        'Comprovante de Endereço': doc_comprovante_var.get(),
        'Carteira de Vacinação': doc_vacinacao_var.get(),
        'Certidão de Nascimento': doc_certidao_var.get(),
        'Ficha de Aproveitamento de Estudos': doc_aproveitamento_var.get(),
        'Relatório Médico': doc_relatorio_var.get(),
        'Declaração de Transferência': doc_transferencia_var.get(),
        'Requerimento de Transferência': doc_req_transferencia_var.get(),
        'Declaração de Matrícula': doc_decl_matricula_var.get(),
        'Boletim Escolar': doc_boletim_var.get(),
        'Certificado de Conclusão': doc_certificado_var.get()
    }

    # Chamar a função para preencher o formulário na planilha Excel
    populate_form(form_data)

    # Exibir uma mensagem de sucesso
    messagebox.showinfo("Sucesso", "Ficha preenchida com sucesso!")

# Criar a interface gráfica do formulário
root = tk.Tk()
root.title("Formulário de Matrícula")

# Campos do formulário
nome_label = tk.Label(root, text="Nome:")
nome_label.grid(row=0, column=0)
nome_entry = tk.Entry(root)
nome_entry.grid(row=0, column=1)

rg_label = tk.Label(root, text="RG:")
rg_label.grid(row=1, column=0)
rg_entry = tk.Entry(root)
rg_entry.grid(row=1, column=1)

cpf_label = tk.Label(root, text="CPF:")
cpf_label.grid(row=2, column=0)
cpf_entry = tk.Entry(root)
cpf_entry.grid(row=2, column=1)

ra_label = tk.Label(root, text="RA:")
ra_label.grid(row=3, column=0)
ra_entry = tk.Entry(root)
ra_entry.grid(row=3, column=1)

estado_civil_label = tk.Label(root, text="Estado Civil:")
estado_civil_label.grid(row=4, column=0)
estado_civil_entry = tk.Entry(root)
estado_civil_entry.grid(row=4, column=1)

cor_label = tk.Label(root, text="Cor/Raça:")
cor_label.grid(row=5, column=0)
cor_var = tk.StringVar(value="Branco")
cor_options = ["Branco", "Preto", "Pardo", "Amarelo", "Indígena", "Outra"]
cor_menu = tk.OptionMenu(root, cor_var, *cor_options)
cor_menu.grid(row=5, column=1)

nome_mae_label = tk.Label(root, text="Nome da Mãe:")
nome_mae_label.grid(row=6, column=0)
nome_mae_entry = tk.Entry(root)
nome_mae_entry.grid(row=6, column=1)

gemeo_label = tk.Label(root, text="Gêmeo:")
gemeo_label.grid(row=7, column=0)
gemeo_var = tk.StringVar(value="Não")
gemeo_sim = tk.Radiobutton(root, text="Sim", variable=gemeo_var, value="Sim")
gemeo_sim.grid(row=7, column=1)
gemeo_nao = tk.Radiobutton(root, text="Não", variable=gemeo_var, value="Não")
gemeo_nao.grid(row=7, column=2)

nascimento_label = tk.Label(root, text="Nascimento:")
nascimento_label.grid(row=8, column=0)
nascimento_entry = tk.Entry(root)
nascimento_entry.grid(row=8, column=1)

municipio_label = tk.Label(root, text="Município:")
municipio_label.grid(row=9, column=0)
municipio_entry = tk.Entry(root)
municipio_entry.grid(row=9, column=1)

uf_label = tk.Label(root, text="UF:")
uf_label.grid(row=10, column=0)
uf_entry = tk.Entry(root)
uf_entry.grid(row=10, column=1)

pais_label = tk.Label(root, text="País:")
pais_label.grid(row=11, column=0)
pais_entry = tk.Entry(root)
pais_entry.grid(row=11, column=1)

itinerario_label = tk.Label(root, text="Opção de Itinerário:")
itinerario_label.grid(row=12, column=0)
itinerario_var = tk.StringVar(value="Ciências Naturais/Matemática")
itinerario_options = ["Ciências Naturais/Matemática", "Linguagens/Sociais"]
itinerario_menu = tk.OptionMenu(root, itinerario_var, *itinerario_options)
itinerario_menu.grid(row=12, column=1)

endereco_label = tk.Label(root, text="Endereço:")
endereco_label.grid(row=13, column=0)
endereco_entry = tk.Entry(root)
endereco_entry.grid(row=13, column=1)

numero_label = tk.Label(root, text="Número:")
numero_label.grid(row=14, column=0)
numero_entry = tk.Entry(root)
numero_entry.grid(row=14, column=1)

bairro_label = tk.Label(root, text="Bairro:")
bairro_label.grid(row=15, column=0)
bairro_entry = tk.Entry(root)
bairro_entry.grid(row=15, column=1)

zona_label = tk.Label(root, text="Zona Urbana/Rural:")
zona_label.grid(row=16, column=0)
zona_var = tk.StringVar(value="Urbana")
zona_urbana = tk.Radiobutton(root, text="Urbana", variable=zona_var, value="Urbana")
zona_urbana.grid(row=16, column=1)
zona_rural = tk.Radiobutton(root, text="Rural", variable=zona_var, value="Rural")
zona_rural.grid(row=16, column=2)

cep_label = tk.Label(root, text="CEP:")
cep_label.grid(row=17, column=0)
cep_entry = tk.Entry(root)
cep_entry.grid(row=17, column=1)

cidade_label = tk.Label(root, text="Cidade:")
cidade_label.grid(row=18, column=0)
cidade_entry = tk.Entry(root)
cidade_entry.grid(row=18, column=1)

uf_cidade_label = tk.Label(root, text="UF_Cidade:")
uf_cidade_label.grid(row=19, column=0)
uf_cidade_entry = tk.Entry(root)
uf_cidade_entry.grid(row=19, column=1)

telefone_celular_label = tk.Label(root, text="Telefone Celular:")
telefone_celular_label.grid(row=20, column=0)
telefone_celular_entry = tk.Entry(root)
telefone_celular_entry.grid(row=20, column=1)

telefone_recado_label = tk.Label(root, text="Telefone Recado:")
telefone_recado_label.grid(row=21, column=0)
telefone_recado_entry = tk.Entry(root)
telefone_recado_entry.grid(row=21, column=1)

nivel_label = tk.Label(root, text="Requer Matrícula no:")
nivel_label.grid(row=22, column=0)
nivel_var = tk.StringVar(value="Ensino Fundamental")
nivel_options = ["Ensino Fundamental", "Ensino Médio", "EJA"]
nivel_menu = tk.OptionMenu(root, nivel_var, *nivel_options)
nivel_menu.grid(row=22, column=1)

serie_label = tk.Label(root, text="Termo/Série:")
serie_label.grid(row=23, column=0)
serie_var = tk.StringVar(value="1º Termo")
serie_options = ["1º Termo", "2º Termo", "3º Termo", "4º Termo", "1ª Série", "2ª Série", "3ª Série", "4ª Série"]
serie_menu = tk.OptionMenu(root, serie_var, *serie_options)
serie_menu.grid(row=23, column=1)

ensino_religioso_label = tk.Label(root, text="Ensino Religioso:")
ensino_religioso_label.grid(row=24, column=0)
ensino_religioso_var = tk.StringVar(value="Sim")
ensino_religioso_sim = tk.Radiobutton(root, text="Sim", variable=ensino_religioso_var, value="Sim")
ensino_religioso_sim.grid(row=24, column=1)
ensino_religioso_nao = tk.Radiobutton(root, text="Não", variable=ensino_religioso_var, value="Não")
ensino_religioso_nao.grid(row=24, column=2)

estudou_ue_label = tk.Label(root, text="Estudou nesta U.E.:")
estudou_ue_label.grid(row=25, column=0)
estudou_ue_var = tk.StringVar(value="Sim")
estudou_ue_sim = tk.Radiobutton(root, text="Sim", variable=estudou_ue_var, value="Sim")
estudou_ue_sim.grid(row=25, column=1)
estudou_ue_nao = tk.Radiobutton(root, text="Não", variable=estudou_ue_var, value="Não")
estudou_ue_nao.grid(row=25, column=2)

aproveitamento_label = tk.Label(root, text="Aproveitamento de Estudos:")
aproveitamento_label.grid(row=26, column=0)
aproveitamento_var = tk.StringVar(value="Sim")
aproveitamento_sim = tk.Radiobutton(root, text="Sim", variable=aproveitamento_var, value="Sim")
aproveitamento_sim.grid(row=26, column=1)
aproveitamento_nao = tk.Radiobutton(root, text="Não", variable=aproveitamento_var, value="Não")
aproveitamento_nao.grid(row=26, column=2)

necessidades_label = tk.Label(root, text="Portador de necessidades ou PCD:")
necessidades_label.grid(row=27, column=0)
necessidades_var = tk.StringVar(value="Não")
necessidades_sim = tk.Radiobutton(root, text="Sim", variable=necessidades_var, value="Sim")
necessidades_sim.grid(row=27, column=1)
necessidades_nao = tk.Radiobutton(root, text="Não", variable=necessidades_var, value="Não")
necessidades_nao.grid(row=27, column=2)

necessidade_qual_label = tk.Label(root, text="Se sim, qual:")
necessidade_qual_label.grid(row=28, column=0)
necessidade_qual_entry = tk.Entry(root)
necessidade_qual_entry.grid(row=28, column=1)

doc_rg_label = tk.Label(root, text="RG:")
doc_rg_label.grid(row=1, column=2)
doc_rg_var = tk.IntVar()
doc_rg_check = tk.Checkbutton(root, variable=doc_rg_var)
doc_rg_check.grid(row=1, column=3)

doc_cpf_label = tk.Label(root, text="CPF:")
doc_cpf_label.grid(row=2, column=2)
doc_cpf_var = tk.IntVar()
doc_cpf_check = tk.Checkbutton(root, variable=doc_cpf_var)
doc_cpf_check.grid(row=2, column=3)

doc_foto_label = tk.Label(root, text="Foto:")
doc_foto_label.grid(row=3, column=2)
doc_foto_var = tk.IntVar()
doc_foto_check = tk.Checkbutton(root, variable=doc_foto_var)
doc_foto_check.grid(row=3, column=3)

doc_requerimento_label = tk.Label(root, text="Requerimento Disp. Ed, Física:")
doc_requerimento_label.grid(row=4, column=2)
doc_requerimento_var = tk.IntVar()
doc_requerimento_check = tk.Checkbutton(root, variable=doc_requerimento_var)
doc_requerimento_check.grid(row=4, column=3)

doc_historico_label = tk.Label(root, text="Histórico Escolar Ensino Fundamental:")
doc_historico_label.grid(row=5, column=2)
doc_historico_var = tk.IntVar()
doc_historico_check = tk.Checkbutton(root, variable=doc_historico_var)
doc_historico_check.grid(row=5, column=3)

doc_comprovante_label = tk.Label(root, text="Histórico Ensino Médio:")
doc_comprovante_label.grid(row=6, column=2)
doc_comprovante_var = tk.IntVar()
doc_comprovante_check = tk.Checkbutton(root, variable=doc_comprovante_var)
doc_comprovante_check.grid(row=6, column=3)

doc_vacinacao_label = tk.Label(root, text="Comprovante de Residência:")
doc_vacinacao_label.grid(row=7, column=2)
doc_vacinacao_var = tk.IntVar()
doc_vacinacao_check = tk.Checkbutton(root, variable=doc_vacinacao_var)
doc_vacinacao_check.grid(row=7, column=3)

doc_certidao_label = tk.Label(root, text="Outros:")
doc_certidao_label.grid(row=8, column=2)
doc_certidao_var = tk.IntVar()
doc_certidao_check = tk.Checkbutton(root, variable=doc_certidao_var)
doc_certidao_check.grid(row=8, column=3)

doc_aproveitamento_label = tk.Label(root, text="Certidão de Nascimento:")
doc_aproveitamento_label.grid(row=9, column=2)
doc_aproveitamento_var = tk.IntVar()
doc_aproveitamento_check = tk.Checkbutton(root, variable=doc_aproveitamento_var)
doc_aproveitamento_check.grid(row=9, column=3)

doc_relatorio_label = tk.Label(root, text="Reservista:")
doc_relatorio_label.grid(row=10, column=2)
doc_relatorio_var = tk.IntVar()
doc_relatorio_check = tk.Checkbutton(root, variable=doc_relatorio_var)
doc_relatorio_check.grid(row=10, column=3)

doc_transferencia_label = tk.Label(root, text="Título de Eleitor:")
doc_transferencia_label.grid(row=11, column=2)
doc_transferencia_var = tk.IntVar()
doc_transferencia_check = tk.Checkbutton(root, variable=doc_transferencia_var)
doc_transferencia_check.grid(row=11, column=3)

doc_req_transferencia_label = tk.Label(root, text="Carteira de Vacinação:")
doc_req_transferencia_label.grid(row=12, column=2)
doc_req_transferencia_var = tk.IntVar()
doc_req_transferencia_check = tk.Checkbutton(root, variable=doc_req_transferencia_var)
doc_req_transferencia_check.grid(row=12, column=3)

doc_decl_matricula_label = tk.Label(root, text="Atestado de Eliminação de Disciplina:")
doc_decl_matricula_label.grid(row=13, column=2)
doc_decl_matricula_var = tk.IntVar()
doc_decl_matricula_check = tk.Checkbutton(root, variable=doc_decl_matricula_var)
doc_decl_matricula_check.grid(row=13, column=3)

doc_boletim_label = tk.Label(root, text="Declaração de Transferência:")
doc_boletim_label.grid(row=14, column=2)
doc_boletim_var = tk.IntVar()
doc_boletim_check = tk.Checkbutton(root, variable=doc_boletim_var)
doc_boletim_check.grid(row=14, column=3)

doc_certificado_label = tk.Label(root, text="Certificado de Conclusão:")
doc_certificado_label.grid(row=15, column=2)
doc_certificado_var = tk.IntVar()
doc_certificado_check = tk.Checkbutton(root, variable=doc_certificado_var)
doc_certificado_check.grid(row=15, column=3)

# Botão de envio
enviar_button = tk.Button(root, text="Enviar", command=enviar)
enviar_button.grid(row=23, column=4, columnspan=2)

# Executar a aplicação
root.mainloop()



def main_window():
    window = tk.Tk()
    window.title("Sistema de Matrícula")
    window.geometry("950x700")

    def on_search_click():
        passport = passport_entry.get().strip().upper()
        if passport:
            # Supondo que você tenha as funções authenticate, build, search_passport, extract_site_data, e update_excel definidas corretamente
            creds = authenticate()
            service = build('drive', 'v3', credentials=creds)
            result = search_passport(passport, service)
            
            if isinstance(result, dict):  # Resultado encontrado
                ra_number = result.get('RA')
                site_data = extract_site_data(ra_number)
                update_excel(result, site_data)
                messagebox.showinfo("Resultado da Pesquisa", f"Passaporte {passport} encontrado e dados preenchidos!")
            else:  # Passaporte não encontrado
                messagebox.showwarning("Resultado da Pesquisa", result)
        else:
            messagebox.showwarning("Entrada inválida", "Por favor, insira um número de passaporte válido.")

    def on_form_click():
        window.withdraw()  # Esconde a janela principal
        root.deiconify()  # Exibe a janela do formulário

    passport_label = tk.Label(window, text="Número do Passaporte:")
    passport_label.pack(pady=10)
    passport_entry = tk.Entry(window)
    passport_entry.pack(pady=10)

    search_button = tk.Button(window, text="Buscar e Preencher Planilha", command=on_search_click)
    search_button.pack(pady=20)

    form_button = tk.Button(window, text="Preencher Ficha de Matrícula", command=on_form_click)
    form_button.pack(pady=20)

    # Carrega e exibe o logotipo
    if os.path.exists("LOGOTIPO_CEEJA.png"):
        logo = tk.PhotoImage(file="LOGOTIPO_CEEJA.png")
        logo_label = tk.Label(window, image=logo)
        logo_label.image = logo  # Mantém uma referência do logo para exibição
        logo_label.pack(pady=10)

    window.mainloop()

if __name__ == "__main__":
    main_window()