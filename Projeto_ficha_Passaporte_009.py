import os
import io
import tkinter as tk
from tkinter import messagebox
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import warnings
import time
import subprocess
import json
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Configurações
SCOPES = ['https://www.googleapis.com/auth/drive.readonly']
SPREADSHEET_IDS = {
    'F': '10_beB24nkw37BsWQaCrS2lrFspcAarA1',
    'M': '1hGsycXU1JM-saNrluhbHumtsxj6vMRFm'
}
RANGES = {
    'F': [
        '2024_Matriculas',
        'EF_Geral_Alfabetica_31122023'
    ],
    'M': [
        '2024_MATRICULAS',
        'EM_Geral_Alfabetica_31122023'
    ]
}

def authenticate():
    creds = None
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        with open('token.json', 'w') as token:
            token.write(creds.to_json())
    return creds

def download_file(service, file_id):
    request = service.files().get_media(fileId=file_id)
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while not done:
        status, done = downloader.next_chunk()
        print(f"Download {int(status.progress() * 100)}%.")
    fh.seek(0)
    return fh

def search_passport(passport, service):
    prefix = passport[0]  # 'F' ou 'M'
    file_id = SPREADSHEET_IDS.get(prefix)
    sheet_names = RANGES.get(prefix)

    if not file_id or not sheet_names:
        return "Planilha não encontrada para o formato do passaporte."

    file_handle = download_file(service, file_id)
    
    for sheet_name in sheet_names:
        print(f"Procurando na aba {sheet_name}...")
        # Ignore warnings about date parsing issues
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            df = pd.read_excel(file_handle, sheet_name=sheet_name, engine='openpyxl')

        if 'Unnamed: 1' in df.columns:
            passaporte_col = df.columns.get_loc('Unnamed: 1')
            nome_col = df.columns.get_loc('Unnamed: 2')
            ra_col = df.columns.get_loc('Unnamed: 3')
            rg_col = df.columns.get_loc('Unnamed: 4')
            telefone_col = df.columns.get_loc('Unnamed: 5')
            cidade_col = df.columns.get_loc('Unnamed: 6')
            serie_concluida_col = df.columns.get_loc('Unnamed: 9')
            serie_matricular_col = df.columns.get_loc('Unnamed: 10')

            if prefix == 'M':
                prof_disciplina_col = df.columns.get_loc('Unnamed: 23')
                observacao_col = df.columns.get_loc('Unnamed: 34')
            else:  # prefix == 'F'
                prof_disciplina_col = df.columns.get_loc('Unnamed: 18')
                observacao_col = df.columns.get_loc('Unnamed: 25')

            # Limpeza e comparação dos dados
            df.iloc[:, passaporte_col] = df.iloc[:, passaporte_col].astype(str).str.strip()  # Remove espaços extras
            result_row = df[df.iloc[:, passaporte_col] == passport]
            if not result_row.empty:
                # Formata os resultados para exibição
                result = {
                    'Passaporte': result_row.iloc[0, passaporte_col],
                    'Nome': result_row.iloc[0, nome_col],
                    'RA': result_row.iloc[0, ra_col],
                    'RG': result_row.iloc[0, rg_col],
                    'Telefone': result_row.iloc[0, telefone_col],
                    'Cidade': result_row.iloc[0, cidade_col],
                    'Série Concluída': result_row.iloc[0, serie_concluida_col],
                    'Série a ser Matriculado': result_row.iloc[0, serie_matricular_col],
                    'Prof. 1ª Disciplina': result_row.iloc[0, prof_disciplina_col],
                    'Observação': result_row.iloc[0, observacao_col]
                }
                return result

    return "Passaporte não encontrado."

def extract_site_data(ra_number):
    driver_path = 'C:/Users/e497976a/Documents/aut_passap_plani/chromedriver.exe'
    service = Service(executable_path=driver_path)
    driver = webdriver.Chrome(service=service)

    # Carregar dados de login do arquivo JSON
    with open('login_credentials.json') as f:
        credentials = json.load(f)
    username = credentials['username']
    password = credentials['password']
    
    driver.get('https://sed.educacao.sp.gov.br/NCA/FichaAluno/Index')
    time.sleep(3)

    login_field = driver.find_element(By.XPATH, '//*[@id="name"]')
    login_field.send_keys(username)
    password_field = driver.find_element(By.XPATH, '//*[@id="senha"]')
    password_field.send_keys(password)
    time.sleep(3)
    login_button = driver.find_element(By.XPATH, '//*[@id="botaoEntrar"]')
    login_button.click()
    time.sleep(3)
    search_box = driver.find_element(By.XPATH, '//*[@id="decorMenuFilterTxt"]')
    search_box.send_keys('ficha do aluno')
    search_box.send_keys(Keys.ENTER)
    time.sleep(3)
    
    wait = WebDriverWait(driver, 10)
    dropdown_menu = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="Pesquisa"]/div[1]/div[2]/div/div/button/div/div/div')))
    dropdown_menu.click()
    ra_option = wait.until(EC.visibility_of_element_located((By.XPATH, '//*[@id="bs-select-1-1"]/span')))
    ra_option.click()
    search_box_RAaluno = driver.find_element(By.XPATH, '//*[@id="txtRa"]')

    # Converte ra_number para string e remove '-' e o último dígito
    clean_ra_number = str(ra_number).replace('-', '')[:-1] 
    search_box_RAaluno.send_keys(clean_ra_number)
    search_box_Pesquisar = driver.find_element(By.XPATH, '//*[@id="btnPesquisar"]')
    search_box_Pesquisar.click()
    time.sleep(3)
    
    try:
        search_botao_editar = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'i.icone-tabela-editar[title="Editar"]')))
        search_botao_editar.click()
    except Exception as e:
        print(f"Erro ao encontrar o botão 'Editar': {e}")

    time.sleep(3)
    
    site_data = {}
    try:
        endereco_element = driver.find_element(By.XPATH, '//*[@id="Endereco"]')
        site_data['Endereco'] = endereco_element.get_attribute('value')
    except Exception as e:
        print(f'Erro ao encontrar o campo de endereço: {e}')
    try:
        endereco_nr_element = driver.find_element(By.XPATH, '//*[@id="EnderecoNR"]')
        site_data['EnderecoNR'] = endereco_nr_element.get_attribute('value')
    except Exception as e:
        print(f'Erro ao encontrar o número da casa: {e}')
    try:
        cep_element = driver.find_element(By.CSS_SELECTOR, 'input.form-control.cep')
        site_data['CEP'] = cep_element.get_attribute('value')
    except Exception as e:
        print(f'Erro ao encontrar o campo de CEP: {e}')
    try:
        dt_nascimento_element = driver.find_element(By.XPATH, '//*[@id="DtNascimento"]')
        site_data['Data de Nascimento'] = dt_nascimento_element.get_attribute('value')
    except Exception as e:
        print(f'Erro ao encontrar o campo de Data de Nascimento: {e}')
    
    driver.quit()
    return site_data

def update_excel(result, site_data):
    # Abre o arquivo Excel
    wb = load_workbook('Passaporte_Geral_2024.xlsx')
    ws = wb.active

    # Atualizar com dados da pesquisa
    ws['K3'] = result.get('Passaporte', '')
    ws['C3'] = result.get('Nome', '')
    ws['J4'] = result.get('RA', '')
    ws['G4'] = result.get('RG', '')
    ws['C6'] = result.get('Telefone', '')
    ws['I5'] = result.get('Cidade', '')
    ws['D25'] = result.get('Série Concluída', '')
    ws['G25'] = result.get('Série a ser Matriculado', '')
    ws['D26'] = result.get('Observação', '')

    # Atualizar com dados extraídos do site
    if site_data:
        endereco_completo = f"{site_data.get('Endereco', '')}, {site_data.get('EnderecoNR', '')}"
        ws['C5'] = endereco_completo
        ws['L5'] = site_data.get('CEP', '')
        ws['D4'] = site_data.get('Data de Nascimento', '')

    wb.save('Passaporte_Geral_2024.xlsx')

    # Abrir o arquivo em tela
    try:
        if os.name == 'nt':  # Windows
            os.startfile('Passaporte_Geral_2024.xlsx')
        elif os.name == 'posix':  # macOS ou Linux
            subprocess.call(('open', 'Passaporte_Geral_2024.xlsx'))
    except Exception as e:
        print(f"Erro ao abrir o arquivo: {e}")

import tkinter as tk
from tkinter import messagebox, PhotoImage
from openpyxl import load_workbook
import os

def clear_form(ws):
    # Limpar campos específicos
    ws['C6'] = ''
    ws['B7'] = ''
    ws['F7'] = ''
    ws['M7'] = ''
    ws['C8'] = ''

    #Gêmos
    ws['O9'] = 'Sim( )'
    ws['P9'] = 'Não( )'
    
    # Limpar cor/raça
    ws['G8'] = '( )'
    ws['I8'] = '( )'
    ws['K8'] = '( )'
    ws['M8'] = '( )'
    ws['O8'] = '( )'
    ws['Q8'] = '( )'
    
    ws['G9'] = ''
    ws['C9'] = ''
    ws['C10'] = ''
    ws['G10'] = ''
    ws['O10'] = ''
    ws['Q10'] = ''
    
    ws['J14'] = ''
    ws['K13'] = '( )'
    ws['Q13'] = '( )'
    
    ws['B19'] = ''
    ws['O19'] = ''
    ws['C20'] = ''
    ws['M20'] = '( )'
    ws['O20'] = '( )'
    
    ws['B21'] = ''
    ws['F21'] = ''
    ws['M21'] = ''
    ws['F22'] = ''
    ws['M22'] = ''
    
    ws['I25'] = '( )'
    ws['K25'] = '( )'
    ws['M25'] = '( )'
    ws['O25'] = '( )'
    ws['Q25'] = '( )'
    
    ws['I26'] = '( )'
    ws['K26'] = '( )'
    ws['M26'] = '( )'
    ws['O26'] = '( )'
    
    ws['J30'] = '( )'
    ws['L30'] = '( )'
    
    ws['K15'] = '( )'
    ws['M15'] = '( )'
    ws['M16'] = '( )'
    ws['K17'] = '( )'
    ws['M17'] = '( )'

    ws['I31'] = 'Não( )'
    ws['I32'] = 'Sim( )'
    ws['I33'] = '( )'
    ws['I34'] = '( )'
    ws['I35'] = '( )'
    ws['I37'] = '( )'
    ws['I38'] = '( )'

    #limpar campos de documentos entregues
    ws['A40'] = '( )'
    ws['A41'] = '( )'
    ws['A42'] = '( )'
    ws['A43'] = '( )'
    ws['D40'] = '( )'
    ws['D41'] = '( )'
    ws['D42'] = '( )'
    ws['D43'] = '( )'
    ws['I40'] = '( )'
    ws['I41'] = '( )'
    ws['I42'] = '( )'
    ws['M40'] = '( )'
    ws['M41'] = '( )'
    ws['M42'] = '( )'
    ws['A45'] = '( )'
    
def populate_form(form_data):
    # Carregar a planilha Excel
    wb = load_workbook('FICHA_DE_MATRÍCULA_2024.xlsx')
    ws = wb.active

    clear_form(ws)

    # Preencher as células com base nos dados do formulário
    ws['C6'] = form_data.get('Nome', '')
    ws['B7'] = form_data.get('RG', '')
    ws['F7'] = form_data.get('CPF', '')
    ws['M7'] = form_data.get('RA', '')
    ws['C8'] = form_data.get('Estado Civil', '')

    # Cor/raça
    cor_raca = form_data.get('Cor/raça', '')
    ws['G8'] = '(X)' if cor_raca == 'Branco' else '( )'
    ws['I8'] = '(X)' if cor_raca == 'Preto' else '( )'
    ws['K8'] = '(X)' if cor_raca == 'Pardo' else '( )'
    ws['M8'] = '(X)' if cor_raca == 'Amarelo' else '( )'
    ws['O8'] = '(X)' if cor_raca == 'Indígena' else '( )'
    ws['Q8'] = '(X)' if cor_raca == 'Outra' else '( )'
    
    # Outras informações do formulário
    ws['G9'] = form_data.get('Nome da Mãe', '')
   
    #Gêmeo
    gemeo_sim = form_data.get('Gêmeo', '( )')
    
    ws['O9'] = '(x)sim' if gemeo_sim == "Sim" else '( )'
    ws['P9'] = '(x)não' if gemeo_sim == "Não" else '( )'

    ws['C10'] = form_data.get('Nascimento', '')
    ws['G10'] = form_data.get('Município', '')
    ws['O10'] = form_data.get('UF', '')
    ws['Q10'] = form_data.get('País', '')

    # Opção de Itinerário
    if form_data.get('Opção de Itinerário') == 'Ciências Naturais/Matemática':
        ws['K13'] = '(X)'
        ws['Q13'] = '( )'
    else:
        ws['Q13'] = '(X)'
        ws['k13'] = '( )'
    
    # Outros campos preenchidos
    ws['B19'] = form_data.get('Endereço', '')
    ws['O19'] = form_data.get('Número', '')
    ws['C20'] = form_data.get('Bairro', '')
    
    # Zona Urbana ou Rural
  # Preencher as células com base nos dados do formulário
# Zona Urbana ou Rural
    if form_data.get('Urbana/Rural') == 'Urbana':
        ws['M20'] = '( X )'
        ws['O20'] = '(  )'
    elif form_data.get('Urbana/Rural') == 'Rural':
        ws['M20'] = '(  )'
        ws['O20'] = '( X )'
    else:
        ws['M20'] = '(  )'
        ws['O20'] = '(  )'


    ws['B21'] = form_data.get('CEP', '')
    ws['F21'] = form_data.get('Cidade', '')
    ws['M21'] = form_data.get('UF_Cidade', '')
    ws['F22'] = form_data.get('Telefone Celular', '')
    ws['M22'] = form_data.get('Telefone Recado', '')

    # Preenchimento adicional baseado no nível de ensino e série/termo
    nivel_ensino = form_data.get('Requer Matrícula no', '')

    if nivel_ensino == 'Ensino Fundamental':
        ws['I25'] = '(X)'  # Marca "Ensino Fundamental"

        termo = form_data.get('Termo/Série', '')
        if termo == '1º Termo':
            ws['K25'] = '(X)'
        elif termo == '2º Termo':
            ws['M25'] = '(X)'
        elif termo == '3º Termo':
            ws['O25'] = '(X)'
        elif termo == '4º Termo':
            ws['Q25'] = '(X)'

    elif nivel_ensino == 'Ensino Médio':
        ws['I26'] = '(X)'  # Marca "Ensino Médio"

        serie = form_data.get('Termo/Série', '')
        if serie == '1ª Série':
            ws['K26'] = '(X)'
        elif serie == '2ª Série':
            ws['M26'] = '(X)'
        elif serie == '3ª Série':
            ws['O26'] = '(X)'

    if form_data.get('Ensino Religioso') == 'Sim':
        ws['J30'] = '(X)'
    else:
        ws['L30'] = '(X)'

    # Já estudou nesta Unidade?
    if form_data.get('Estudou nesta U.E.') == 'Sim':
        ws['K15'] = '(X)'
    else:
        ws['M15'] = '(X)'

    # Aproveitamento de Estudos
    if form_data.get('Aproveitamento de Estudos') == 'Sim':
        ws['K16'] = '(X)'
    else:
        ws['M16'] = '(X)'

    # Portador de necessidades ou PCD
    if form_data.get('Portador de necessidades ou PCD') == 'Sim':
        ws['K17'] = '(X)'
    else:
        ws['M17'] = '(X)'

    # Necessidade Especial
    ws['J14'] = form_data.get('Se sim, qual', '')
# Preencher os campos de documentos entregues

    if form_data['Doc_RG']:
        ws['A40'] = '(X)'
    if form_data['Doc_CPF']:
        ws['A41'] = '(X)'
    if form_data['Foto']:
        ws['A42'] = '(X)'
    if form_data['Requerimento de Dispensa de Ed. Física']:
        ws['A43'] = '(X)'
    if form_data['Histórico Escolar EF']:
        ws['D40'] = '(X)'
    if form_data['Histórico Escolar EM']:
        ws['D41'] = '(X)'
    if form_data['Comprovante de Residência']:
        ws['D42'] = '(X)'
    if form_data['Outros']:
        ws['D43'] = '(X)'
    if form_data['Certidão de Nascimento/Casamento']:
        ws['I40'] = '(X)'
    if form_data['Reservista']:
        ws['I41'] = '(X)'
    if form_data['Título de Eleitor/TRE']:
        ws['I42'] = '(X)'
    if form_data['Carteira de Vacinação']:
        ws['M40'] = '(X)' 
    if form_data['Atestado de Eliminação de Disciplinas']:
        ws['M41'] = '(X)'
    if form_data['Declaração de Transferência']:
        ws['M42'] = '(X)'

        # Salvar o arquivo Excel
    wb.save('FICHA_DE_MATRÍCULA_2024.xlsx')
    messagebox.showinfo("Sucesso", "Ficha de matrícula preenchida com sucesso!")
def show_form():
    def submit_form():
    # Coletar os dados do formulário
        form_data = {
            "Nome": nome_entry.get(),
            "RG": rg_entry.get(),
            "CPF": cpf_entry.get(),
            "RA": ra_entry.get(),
            "Estado Civil": estado_civil_entry.get(),
            "Cor/raça": cor_raca_var.get(),
            "Nome da Mãe": nome_mae_entry.get(),
            "Gêmeo" : gemeo_sim_var.get(),
            "Nascimento": nascimento_entry.get(),
            "Município": municipio_entry.get(),
            "UF": uf_entry.get(),
            "País": pais_entry.get(),
            "Opção de Itinerário": itinerario_var.get(),
            "Endereço": endereco_entry.get(),
            "Número": numero_entry.get(),
            "Bairro": bairro_entry.get(),
            "Urbana/Rural": zona_var.get(),
            "CEP": cep_entry.get(),
            "Cidade": cidade_entry.get(),
            "UF_Cidade": uf_cidade_entry.get(),
            "Telefone Celular": celular_entry.get(),
            "Telefone Recado": recado_entry.get(),
            "Requer Matrícula no": nivel_ensino_var.get(),
            "Termo/Série": termo_serie_var.get(),
            "Ensino Religioso": religio_var.get(),
            "Estudou nesta U.E.": estudou_var.get(),
            "Aproveitamento de Estudos": aproveitamento_var.get(),
            "Portador de necessidades ou PCD": pcd_var.get(),
            "Se sim, qual?": qual_pcd_entry.get(),
            'Doc_RG': Doc_RG_var.get(),  # Correção: obtendo o valor do BooleanVar
            'Doc_CPF': Doc_CPF_var.get(),  # Correção: obtendo o valor do BooleanVar
            'Foto': foto_var.get(),  # Correção: obtendo o valor do BooleanVar
            'Requerimento de Dispensa de Ed. Física': req_disp_var.get(),  # Correção: obtendo o valor do BooleanVar
            'Histórico Escolar EF': historico_ef_var.get(),  # Correção: obtendo o valor do BooleanVar
            'Histórico Escolar EM': historico_em_var.get(),  # Correção: obtendo o valor do BooleanVar
            'Comprovante de Residência': comprovante_residencia_var.get(),  # Correção: obtendo o valor do BooleanVar
            'Outros': outros_var.get(),  # Correção: obtendo o valor do BooleanVar
            'Certidão de Nascimento/Casamento': cert_nasc_var.get(),  # Correção: obtendo o valor do BooleanVar
            'Reservista': reservista_var.get(),  # Correção: obtendo o valor do BooleanVar
            'Título de Eleitor/TRE': titulo_eleitor_var.get(),  # Correção: obtendo o valor do BooleanVar
            'Carteira de Vacinação': vacina_var.get(),  # Correção: obtendo o valor do BooleanVar
            'Atestado de Eliminação de Disciplinas': atest_elim_var.get(),  # Correção: obtendo o valor do BooleanVar
            'Declaração de Transferência': declaracao_transf_var.get(),  # Correção: obtendo o valor do BooleanVar
     }

        print("Form Data:", form_data)  # Verifique os valores aqui
    
        populate_form(form_data)
    
    # Criação da janela do formulário
    root = tk.Tk()
    root.title("Formulário de Matrícula")
    
    # Criação dos campos do formulário
    tk.Label(root, text="Nome").grid(row=0, column=0)
    nome_entry = tk.Entry(root)
    nome_entry.grid(row=0, column=1)

    tk.Label(root, text="RG").grid(row=1, column=0)
    rg_entry = tk.Entry(root)
    rg_entry.grid(row=1, column=1)

    tk.Label(root, text="CPF").grid(row=2, column=0)
    cpf_entry = tk.Entry(root)
    cpf_entry.grid(row=2, column=1)

    tk.Label(root, text="RA").grid(row=3, column=0)
    ra_entry = tk.Entry(root)
    ra_entry.grid(row=3, column=1)

    tk.Label(root, text="Estado Civil").grid(row=4, column=0)
    estado_civil_entry = tk.Entry(root)
    estado_civil_entry.grid(row=4, column=1)
    
    tk.Label(root, text="Cor/raça").grid(row=5, column=0)
    cor_raca_var = tk.StringVar(value='Branco')
    cor_raca_menu = tk.OptionMenu(root, cor_raca_var, 'Branco', 'Preto', 'Pardo', 'Amarelo', 'Indígena', 'Outra')
    cor_raca_menu.grid(row=5, column=1)
    
    tk.Label(root, text="Nome da Mãe").grid(row=6, column=0)
    nome_mae_entry = tk.Entry(root)
    nome_mae_entry.grid(row=6, column=1)
    
    tk.Label(root, text="Gêmeo").grid(row=7, column=0)
    gemeo_sim_var = tk.StringVar(value='Sim')
    gemeo_menu = tk.OptionMenu(root, gemeo_sim_var, 'Sim', 'Não')
    gemeo_menu.grid(row=7, column=1)
    
    tk.Label(root, text="Nascimento").grid(row=8, column=0)
    nascimento_entry = tk.Entry(root)
    nascimento_entry.grid(row=8, column=1)
    
    tk.Label(root, text="Município").grid(row=9, column=0)
    municipio_entry = tk.Entry(root)
    municipio_entry.grid(row=9, column=1)
    
    tk.Label(root, text="UF").grid(row=10, column=0)
    uf_entry = tk.Entry(root)
    uf_entry.grid(row=10, column=1)
    
    tk.Label(root, text="País").grid(row=11, column=0)
    pais_entry = tk.Entry(root)
    pais_entry.grid(row=11, column=1)
    
    tk.Label(root, text="Opção de Itinerário").grid(row=12, column=0)
    itinerario_var = tk.StringVar(value='Ciências Naturais/Matemática')
    itinerario_menu = tk.OptionMenu(root, itinerario_var, 'Ciências Naturais/Matemática', 'Linguagens e Ciências Humanas')
    itinerario_menu.grid(row=12, column=1)
    
    tk.Label(root, text="Endereço").grid(row=13, column=0)
    endereco_entry = tk.Entry(root)
    endereco_entry.grid(row=13, column=1)
    
    tk.Label(root, text="Número").grid(row=14, column=0)
    numero_entry = tk.Entry(root)
    numero_entry.grid(row=14, column=1)
    
    tk.Label(root, text="Bairro").grid(row=15, column=0)
    bairro_entry = tk.Entry(root)
    bairro_entry.grid(row=15, column=1) 

    zona_var = tk.StringVar(value='Urbana')
    tk.Radiobutton(root, text="Urbana", variable=zona_var, value='Urbana').grid(row=16, column=1)
    tk.Radiobutton(root, text="Rural", variable=zona_var, value='Rural').grid(row=16, column=2)
    
    tk.Label(root, text="CEP").grid(row=17, column=0)
    cep_entry = tk.Entry(root)
    cep_entry.grid(row=17, column=1)
    
    tk.Label(root, text="Cidade").grid(row=18, column=0)
    cidade_entry = tk.Entry(root)
    cidade_entry.grid(row=18, column=1)
    
    tk.Label(root, text="UF_Cidade").grid(row=19, column=0)
    uf_cidade_entry = tk.Entry(root)
    uf_cidade_entry.grid(row=19, column=1)
    
    tk.Label(root, text="Telefone Celular").grid(row=20, column=0)
    celular_entry = tk.Entry(root)
    celular_entry.grid(row=20, column=1)
    
    tk.Label(root, text="Telefone Recado").grid(row=21, column=0)
    recado_entry = tk.Entry(root)
    recado_entry.grid(row=21, column=1)
    
    # Adiciona campos para nível de ensino e série/termo
    tk.Label(root, text="Requer Matrícula no").grid(row=22, column=0)
    nivel_ensino_var = tk.StringVar()
    tk.OptionMenu(root, nivel_ensino_var, "Ensino Fundamental", "Ensino Médio").grid(row=22, column=1)

    tk.Label(root, text="Termo/Série").grid(row=23, column=0)
    termo_serie_var = tk.StringVar()
    tk.OptionMenu(root, termo_serie_var, "1º Termo", "2º Termo", "3º Termo", "4º Termo", "1ª Série", "2ª Série", "3ª Série").grid(row=23, column=1)
    tk.Label(root, text="Ensino Religioso").grid(row=24, column=0)
    religio_var = tk.StringVar(value='Sim')
    religio_menu = tk.OptionMenu(root, religio_var, 'Sim', 'Não')
    religio_menu.grid(row=24, column=1)
    
    tk.Label(root, text="Estudou nesta U.E.").grid(row=25, column=0)
    estudou_var = tk.StringVar(value='Sim')
    estudou_menu = tk.OptionMenu(root, estudou_var, 'Sim', 'Não')
    estudou_menu.grid(row=25, column=1)
    
    tk.Label(root, text="Aproveitamento de Estudos").grid(row=1, column=2)
    aproveitamento_var = tk.StringVar(value='Sim')
    aproveitamento_menu = tk.OptionMenu(root, aproveitamento_var, 'Sim', 'Não')
    aproveitamento_menu.grid(row=1, column=3)

    tk.Label(root, text="Portador de Necessidades Especiais?").grid(row=2, column=2)
    pcd_var = tk.StringVar(value='Sim')
    pcd_menu = tk.OptionMenu(root, pcd_var, 'Sim', 'Não')
    pcd_menu.grid(row=2, column=3)

    tk.Label(root, text="Se sim, Qual?").grid(row=4, column=2)
    qual_pcd_entry = tk.Entry(root)
    qual_pcd_entry.grid(row=4, column=3)

    # Adicionando campos de documentos entregues
    tk.Label(root, text="Documentos Entregues:").grid(row=5, column=2, sticky="w")

    Doc_RG_var = tk.BooleanVar()
    tk.Checkbutton(root, text="Doc RG", variable=Doc_RG_var).grid(row=6, column=3, sticky="w")

    Doc_CPF_var = tk.BooleanVar()
    tk.Checkbutton(root, text="Doc CPF", variable=Doc_CPF_var).grid(row=7, column=3, sticky="w")

    foto_var = tk.BooleanVar()
    tk.Checkbutton(root, text="Foto", variable=foto_var).grid(row=8, column=4, sticky="w")

    req_disp_var = tk.BooleanVar()
    tk.Checkbutton(root, text="Requerimento de Dispensa de Ed. Física", variable=req_disp_var).grid(row=8, column=5, sticky="w")

    historico_ef_var = tk.BooleanVar()
    tk.Checkbutton(root, text="Histórico Escolar EF", variable=historico_ef_var).grid(row=9, column=3, sticky="w")

    historico_em_var = tk.BooleanVar()
    tk.Checkbutton(root, text="Histórico Escolar EM", variable=historico_em_var).grid(row=9, column=4, sticky="w")

    comprovante_residencia_var = tk.BooleanVar()
    tk.Checkbutton(root, text="Comprovante de Residência", variable=comprovante_residencia_var).grid(row=9, column=5, sticky="w")

    outros_var = tk.BooleanVar()
    tk.Checkbutton(root, text="Outros", variable=outros_var).grid(row=10, column=3, sticky="w")

    cert_nasc_var = tk.BooleanVar()
    tk.Checkbutton(root, text="Certidão de Nascimento/Casamento", variable=cert_nasc_var).grid(row=10, column=4, sticky="w")

    reservista_var = tk.BooleanVar()
    tk.Checkbutton(root, text="Reservista", variable=reservista_var).grid(row=10, column=5, sticky="w")

    titulo_eleitor_var = tk.BooleanVar()
    tk.Checkbutton(root, text="Título de Eleitor/TRE", variable=titulo_eleitor_var).grid(row=11, column=3, sticky="w")

    vacina_var = tk.BooleanVar()
    tk.Checkbutton(root, text="Carteira de Vacinação", variable=vacina_var).grid(row=11, column=4, sticky="w")

    atest_elim_var = tk.BooleanVar()
    tk.Checkbutton(root, text="Atestado de Eliminação de Disciplinas", variable=atest_elim_var).grid(row=11, column=5, sticky="w")

    declaracao_transf_var = tk.BooleanVar()
    tk.Checkbutton(root, text="Declaração de Transferência", variable=declaracao_transf_var).grid(row=11, column=3, sticky="w")
                   
    
    # Botão de envio do formulário
    tk.Button(root, text="Enviar", command=submit_form).grid(row=28, column=0, columnspan=2)

    # Iniciar o loop principal da interface
    root.mainloop()

def main_window():
    window = tk.Tk()
    window.title("Sistema de Matrícula")
    window.geometry("800x600")

    def on_search_click():
        passport = passport_entry.get().strip().upper()
        if passport:
            # Supondo que você tenha as funções authenticate, build, search_passport, extract_site_data, e update_excel definidas corretamente
            creds = authenticate()
            service = build('drive', 'v3', credentials=creds)
            result = search_passport(passport, service)
            
            if isinstance(result, dict):  # Resultado encontrado
                ra_number = result.get('RA')
                site_data = extract_site_data(ra_number)
                update_excel(result, site_data)
                messagebox.showinfo("Resultado da Pesquisa", f"Passaporte {passport} encontrado e dados preenchidos!")
            else:  # Passaporte não encontrado
                messagebox.showwarning("Resultado da Pesquisa", result)
        else:
            messagebox.showwarning("Entrada inválida", "Por favor, insira um número de passaporte válido.")

    def on_form_click():
        window.withdraw()  # Esconde a janela principal
        show_form()  # Exibe a janela do formulário

    passport_label = tk.Label(window, text="Número do Passaporte:")
    passport_label.pack(pady=10)
    passport_entry = tk.Entry(window)
    passport_entry.pack(pady=10)

    search_button = tk.Button(window, text="Buscar e Preencher Planilha", command=on_search_click)
    search_button.pack(pady=20)

    form_button = tk.Button(window, text="Preencher Ficha de Matrícula", command=on_form_click)
    form_button.pack(pady=20)

    # Carrega e exibe o logotipo
    if os.path.exists("LOGOTIPO_CEEJA.png"):
        logo = PhotoImage(file="LOGOTIPO_CEEJA.png")
        logo_label = tk.Label(window, image=logo)
        logo_label.image = logo  # Mantém uma referência do logo para exibição
        logo_label.pack(pady=10)

    window.mainloop()

if __name__ == "__main__":
    main_window()